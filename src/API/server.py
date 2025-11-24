from flask import Flask, request, jsonify
from flask_cors import CORS
import pandas as pd
import mysql.connector
from mysql.connector import Error as MySQLError
import os
import openpyxl
import bcrypt
import re
from typing import Optional, Any

app = Flask(__name__)
CORS(app)

# Configuration
UPLOAD_FOLDER = './uploads'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# Database Configuration
DB_CONFIG = {
    'host': 'localhost',
    'user': 'root',
    'password': '#Bb108935',
    'database': 'salary_db'
}

# Column mappings
COLUMN_MAP = {
    'cid': 'cid', 
    'เลขบัตรประชาชน': 'cid',
    'ชื่อ': 'name', 
    'เลขที่บัญชี': 'bank_account',
    'เดือน': 'month', 
    'ปี': 'year',
    'เงินเดือน': 'salary',
    'เงินเดือนสุทธิ': 'salary_deductions',
    'รวมรับ': 'total_income',
    'ค่าครองชีพ': 'cola_allowance',
    'ค่าครองชีพ(ตกเบิก)': 'retroactive_cola_allowance',
    'ง/ด(ตกเบิก)': 'retroactive_salary_emp',
    'ง/ด ตกเบิก': 'retroactive_salary_emp',
    'พตส.': 'special_public_health_allowance',
    'ปจต.': 'position_allowance',
    'รายเดือน': 'monthly_allowance',
    'P4P': 'pay_for_performance',
    'โควิด-19': 'covid_risk_pay',
    'เพิ่มโควิด-19': 'covid_risk_pay',
    'เสี่ยงภัยโควิด': 'covid_exposure',
    'เงินกู้สวัสดิการ': 'welfare_loan_received',
    'โอที': 'overtime_pay',
    'บ่าย-ดึก': 'evening_night_shift_pay',
    'OT/OPD': 'ot_outpatient_dept',
    'OT/พบ.': 'ot_professional',
    'OT/ผช.': 'ot_assistant',
    'บ-ด/พบ.': 'shift_professional',
    'บ-ด/ผช.': 'shift_assistant',
    'อื่นๆ': 'other_income',
    'รวมจ่าย': 'total_expense',
    'คงเหลือ': 'net_balance',
    'หักวันลา': 'leave_day_deduction',
    'ภาษี': 'tax_deduction',
    'ภาษี ตกเบิก': 'retroactive_tax_deduction',
    'กบข.': 'gpf_contribution',
    'กบข.ตกเบิก': 'retroactive_gpf_deduction',
    'กบข.เพิ่ม': 'gpf_extra_contribution',
    'ปกสค.': 'social_security_deduction_gov',
    'ประกันสังคม': 'social_security_deduction_emp',
    'กองทุน พกส.': 'phks_provident_fund',
    'ฌกส.': 'funeral_welfare_deduction',
    'สอ.กรม': 'coop_deduction_dept',
    'สอ.สสจ.เลย': 'coop_deduction_phso',
    'สสจ.': 'prov_health_office',
    'กยศ.': 'student_loan_deduction_emp',
    'กยศ': 'student_loan_deduction_emp',
    'ค่าน้ำ': 'water_bill_deduction',
    'ค่าไฟ': 'electricity_bill_deduction',
    'net': 'internet_deduction_emp',
    'ค่าNet': 'internet_deduction_emp',
    'AIA': 'aia_insurance_deduction_emp',
    'ค่าAIA': 'aia_insurance_deduction_emp',
    'ออมสิน': 'gsb_loan_deduction_emp',
    'ออมสินนาอาน': 'gsb_loan_naan',
    'ธนาคารออมสินเลย': 'gsb_loan_loei',
    'ธอส': 'ghb_loan_deduction',
    'กรุงไทย': 'ktb_loan_deduction_emp',
    'ธนาคารกรุงไทย': 'ktb_loan_deduction_emp',
    'เงินกู้ รพ.': 'hospital_loan_deduction',
    'เงินกู้ รพ/ประกันงาน': 'hospital_loan_employment',    
    'การศึกษาบุตร': 'child_education_deduction',
    'ค่ารักษาพยาบาล': 'medical_expense_deduction',
    'ไม่ปฏิบัติเวช': 'no_private_practice_deduction',
}

NUMERIC_COLUMNS = {
    'salary', 'salary_deductions', 'total_income', 'cola_allowance',
    'retroactive_cola_allowance', 'retroactive_salary_emp', 'special_public_health_allowance',
    'position_allowance', 'monthly_allowance', 'pay_for_performance', 'covid_risk_pay',
    'welfare_loan_received', 'overtime_pay', 'evening_night_shift_pay', 'ot_outpatient_dept',
    'ot_professional', 'ot_assistant', 'shift_professional', 'shift_assistant', 'other_income',
    'total_expense', 'net_balance', 'leave_day_deduction', 'tax_deduction',
    'retroactive_tax_deduction', 'gpf_contribution', 'retroactive_gpf_deduction',
    'gpf_extra_contribution', 'social_security_deduction_gov', 'social_security_deduction_emp',
    'phks_provident_fund', 'funeral_welfare_deduction', 'coop_deduction_dept',
    'coop_deduction_phso', 'student_loan_deduction_emp', 'water_bill_deduction',
    'electricity_bill_deduction', 'internet_deduction_emp', 'aia_insurance_deduction_emp',
    'gsb_loan_deduction_emp', 'gsb_loan_naan', 'gsb_loan_loei', 'ghb_loan_deduction',
    'ktb_loan_deduction_emp', 'hospital_loan_deduction','hospital loan_employment', 'child_education_deduction',
    'medical_expense_deduction', 'no_private_practice_deduction'
}

MONTH_MAP = {
    'มกราคม': 1, 'กุมภาพันธ์': 2, 'มีนาคม': 3, 'เมษายน': 4,
    'พฤษภาคม': 5, 'มิถุนายน': 6, 'กรกฎาคม': 7, 'สิงหาคม': 8,
    'กันยายน': 9, 'ตุลาคม': 10, 'พฤศจิกายน': 11, 'ธันวาคม': 12,
    'ม.ค.': 1, 'ก.พ.': 2, 'มี.ค.': 3, 'เม.ย.': 4,
    'พ.ค.': 5, 'มิ.ย.': 6, 'ก.ค.': 7, 'ส.ค.': 8,
    'ก.ย.': 9, 'ต.ค.': 10, 'พ.ย.': 11, 'ธ.ค.': 12
}

MONTH_NAMES = {
    1: 'มกราคม', 2: 'กุมภาพันธ์', 3: 'มีนาคม', 4: 'เมษายน',
    5: 'พฤษภาคม', 6: 'มิถุนายน', 7: 'กรกฎาคม', 8: 'สิงหาคม',
    9: 'กันยายน', 10: 'ตุลาคม', 11: 'พฤศจิกายน', 12: 'ธันวาคม'
}

EXCLUDE_COLUMNS = {"ลำดับที่"}


# ============================================================================
#                          DATABASE UTILITIES
# ============================================================================

def get_db_connection():
    """Get database connection"""
    try:
        conn = mysql.connector.connect(**DB_CONFIG)
        return conn
    except MySQLError as e:
        raise


def close_db_connection(conn):
    """Close database connection safely"""
    try:
        if conn and conn.is_connected():
            conn.close()
    except Exception as e:
        pass

# ============================================================================
#                          PASSWORD UTILITIES
# ============================================================================

def hash_password(password: str) -> str:
    """Hash password using bcrypt"""
    salt = bcrypt.gensalt()
    hashed = bcrypt.hashpw(password.encode('utf-8'), salt)
    return hashed.decode('utf-8')


def verify_password(password: str, hashed_password: str) -> bool:
    """Verify password against bcrypt hash"""
    try:
        return bcrypt.checkpw(
            password.encode('utf-8'), 
            hashed_password.encode('utf-8')
        )
    except Exception as e:
        return False


# ============================================================================
#                          DATA CLEANING UTILITIES
# ============================================================================

def clean_column_name(col: str) -> str:
    """Clean and standardize column names"""
    c = str(col).strip()
    if c.lower() in ['', 'nan', 'none', '#ref!']:
        c = "col_empty"
    
    safe = (c.replace(" ", "_")
             .replace("/", "_")
             .replace("-", "_")
             .replace(".", "_")
             .replace("(", "")
             .replace(")", "")
             .replace("%", "pct")
             .replace("#", "no"))
    return safe[:64]


def clean_value(val: Any) -> str:
    """Clean cell values"""
    if val is None or pd.isna(val):
        return ''
    s = str(val).strip()
    if s.lower() in ['nan', 'none', '#ref!', 'null']:
        return ''
    return s


def convert_to_decimal(value: Any) -> Optional[float]:
    """Convert value to decimal, removing commas and special characters"""
    if value is None or value == '' or pd.isna(value):
        return None
    
    s = str(value).strip()
    s = s.replace(',', '')
    s = re.sub(r'[^\d.-]', '', s)
    
    try:
        return float(s) if s else None
    except:
        return None


def convert_month_to_number(month_value: Any) -> Optional[int]:
    """Convert Thai month name or number to integer"""
    if month_value is None or month_value == '' or pd.isna(month_value):
        return None
    
    try:
        num = int(month_value)
        if 1 <= num <= 12:
            return num
    except:
        pass
    
    month_str = str(month_value).strip()
    return MONTH_MAP.get(month_str, None)


def clean_cid(value: Any) -> Optional[str]:
    """Clean and validate CID (13 digits)"""
    if value is None or value == '' or pd.isna(value):
        return None
    
    # ลบทุกอย่างที่ไม่ใช่ตัวเลข
    cid = re.sub(r'\D', '', str(value))
    
    # ต้องมี 13 หลักพอดี
    if len(cid) == 13:
        return cid
    
    # ถ้าไม่ใช่ 13 หลัก ให้คืนค่าเป็น string ธรรมดา
    return cid if cid else None


def clean_year(value: Any) -> Optional[int]:
    """Clean and validate year - keep as Buddhist Era (พ.ศ.)"""
    if value is None or value == '' or pd.isna(value):
        return None
    
    try:
        year = int(float(str(value).replace(',', '').strip()))
        
        # ✅ เก็บเป็น พ.ศ. ตามที่รับเข้ามา (ไม่แปลงเป็น ค.ศ.)
        # ถ้าเป็นปี ค.ศ. (เช่น 2024) ให้แปลงเป็น พ.ศ.
        if year < 2500:
            year = year + 543
            
        return year
    except:
        return None


def clean_bank_account(value: Any) -> Optional[str]:
    """Clean bank account number"""
    if value is None or value == '' or pd.isna(value):
        return None
    
    try:
        # ลบเครื่องหมาย , - และ space
        account = str(value).replace(',', '').replace('-', '').replace(' ', '').strip()
        # ลบทุกอย่างที่ไม่ใช่ตัวเลข
        account = re.sub(r'\D', '', account)
        return account if account else None
    except:
        return None

# ============================================================================
#                          AUTHENTICATION ROUTES
# ============================================================================

@app.route("/api/register", methods=["POST"])
def register():
    """Register new user"""
    conn = None
    try:
        data = request.get_json()
        cid = data.get('cid', '').strip()
        password = data.get('password', '').strip()
        name = data.get('name', '').strip()
        status = data.get('status', '').strip()
        
        # Validation
        if not all([cid, password, name, status]):
            return jsonify({
                "status": "error",
                "message": "กรุณากรอกข้อมูลให้ครบถ้วน"
            }), 400
        
        # Hash password
        hashed_pwd = hash_password(password)
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Check if CID exists
        cursor.execute("SELECT cid FROM users WHERE cid = %s", (cid,))
        if cursor.fetchone():
            cursor.close()
            return jsonify({
                "status": "error",
                "message": "CID นี้มีในระบบแล้ว"
            }), 400
        
        # Insert new user
        cursor.execute(
            """INSERT INTO users (cid, password, name, status)
               VALUES (%s, %s, %s, %s)""",
            (cid, hashed_pwd, name, status)
        )
        conn.commit()
        cursor.close()
        
        return jsonify({
            "status": "success",
            "message": "ลงทะเบียนสำเร็จ"
        })
        
    except Exception as e:
        print(f"Register error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            "status": "error",
            "message": "เกิดข้อผิดพลาดในการลงทะเบียน"
        }), 500
    finally:
        close_db_connection(conn)


@app.route("/api/login", methods=["POST"])
def login():
    """User login - supports both plain text and bcrypt"""
    conn = None
    try:
        data = request.get_json()
        cid = data.get('cid', '').strip()
        password = data.get('password', '').strip()
        
        if not cid or not password:
            return jsonify({
                "status": "error",
                "message": "กรุณากรอก CID และรหัสผ่าน"
            }), 400
        
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute(
            "SELECT * FROM users WHERE cid = %s",
            (cid,)
        )
        user = cursor.fetchone()
        cursor.close()
        
        if not user:
            return jsonify({
                "status": "error",
                "message": "CID หรือรหัสผ่านไม่ถูกต้อง"
            }), 401
        
        # Check if password is bcrypt hash or plain text
        stored_password = user['password']
        password_valid = False
        
        if stored_password.startswith('$2b$'):
            # Bcrypt hashed password
            password_valid = verify_password(password, stored_password)
        else:
            # Plain text password (temporary)
            password_valid = (password == stored_password)
        
        if not password_valid:
            return jsonify({
                "status": "error",
                "message": "CID หรือรหัสผ่านไม่ถูกต้อง"
            }), 401
        
        return jsonify({
            "status": "success",
            "message": "เข้าสู่ระบบสำเร็จ",
            "user": {
                "cid": user['cid'],
                "name": user['name'],
                "status": user.get('status', '')
            }
        })
            
    except Exception as e:
        print(f"Login error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            "status": "error",
            "message": "เกิดข้อผิดพลาดในการเข้าสู่ระบบ"
        }), 500
    finally:
        close_db_connection(conn)


@app.route("/api/hash-passwords", methods=["POST"])
def hash_existing_passwords():
    """Migration utility: Hash existing plain text passwords"""
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT id, cid, password FROM users")
        users = cursor.fetchall()
        
        updated_count = 0
        skipped_count = 0
        
        for user in users:
            # Skip if already hashed
            if user['password'].startswith('$2b$'):
                skipped_count += 1
                continue
            
            hashed = hash_password(user['password'])
            cursor.execute(
                "UPDATE users SET password = %s WHERE id = %s",
                (hashed, user['id'])
            )
            updated_count += 1
        
        conn.commit()
        cursor.close()
        
        return jsonify({
            "status": "success",
            "message": "เข้ารหัสรหัสผ่านสำเร็จ",
            "updated": updated_count,
            "skipped": skipped_count
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({
            "status": "error",
            "message": str(e)
        }), 500
    finally:
        close_db_connection(conn)


# ============================================================================
#                          USER MANAGEMENT ROUTES
# ============================================================================

@app.route("/api/users", methods=["GET"])
def get_users():
    """Get all users (for admin)"""
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute(
            "SELECT cid, name, status, created_at FROM users ORDER BY created_at DESC"
        )
        users = cursor.fetchall()
        cursor.close()
        
        return jsonify({
            "status": "success",
            "users": users
        })
        
    except Exception as e:
        return jsonify({
            "status": "error",
            "message": str(e)
        }), 500
    finally:
        close_db_connection(conn)


@app.route("/api/users/<cid>", methods=["DELETE"])
def delete_user(cid):
    """Delete user by CID"""
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM users WHERE cid = %s", (cid,))
        
        if cursor.rowcount == 0:
            cursor.close()
            return jsonify({
                "status": "error",
                "message": "ไม่พบผู้ใช้"
            }), 404
        
        conn.commit()
        cursor.close()
        
        return jsonify({
            "status": "success",
            "message": "ลบผู้ใช้สำเร็จ"
        })
        
    except Exception as e:
        return jsonify({
            "status": "error",
            "message": str(e)
        }), 500
    finally:
        close_db_connection(conn)


# ============================================================================
#                          SALARY DATA ROUTES
# ============================================================================

@app.route("/api/salary-data", methods=["GET"])
def get_salary_data():
    """Get salary data with filters"""
    conn = None
    try:
        cid = request.args.get('cid', '')
        name = request.args.get('name', '')
        month = request.args.get('month', '')
        year = request.args.get('year', '')
        employee = request.args.get('employee', '')
        
        # Convert month
        month_number = None
        if month:
            month_number = convert_month_to_number(month)
        
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        query = "SELECT * FROM salary_data WHERE 1=1"
        params = []
        
        if cid:
            query += " AND cid LIKE %s"
            params.append(f"%{cid}%")
        
        if name:
            query += " AND name LIKE %s"
            params.append(f"%{name}%")
            
        if month_number is not None:
            query += " AND month = %s"
            params.append(month_number)
            
        if year:
            query += " AND year = %s"
            params.append(year)
            
        if employee:
            query += " AND employee = %s"
            params.append(employee)
        
        query += """
            ORDER BY 
                CASE 
                    WHEN employee = 'ข้าราชการ' THEN 0 
                    WHEN employee = 'ลูกจ้างเงินเดือน' THEN 1 
                    ELSE 2 
                END,
                id ASC
        """
        
        cursor.execute(query, params)
        results = cursor.fetchall()
        
        # ✅ แปลง BIGINT เป็น string สำหรับ Frontend
        for row in results:
            if row.get('cid'):
                row['cid'] = str(row['cid'])
            if row.get('bank_account'):
                row['bank_account'] = str(row['bank_account'])
        
        cursor.close()
        
        return jsonify({
            "status": "success",
            "data": results,
            "count": len(results)
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({
            "status": "error",
            "message": str(e)
        }), 500
    finally:
        close_db_connection(conn)


@app.route("/api/available-filters", methods=["GET"])
def get_available_filters():
    """Get available months and years for filtering"""
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Get months
        cursor.execute(
            "SELECT DISTINCT month FROM salary_data WHERE month IS NOT NULL ORDER BY month"
        )
        months_data = cursor.fetchall()
        
        available_months = []
        for (month_num,) in months_data:
            if month_num in MONTH_NAMES:
                available_months.append({
                    'value': MONTH_NAMES[month_num],
                    'label': MONTH_NAMES[month_num],
                    'number': month_num
                })
        
        # Get years
        cursor.execute(
            "SELECT DISTINCT year FROM salary_data WHERE year IS NOT NULL ORDER BY year DESC"
        )
        years_data = cursor.fetchall()
        available_years = [str(year[0]) for year in years_data]
        cursor.close()
        
        return jsonify({
            "status": "success",
            "months": available_months,
            "years": available_years
        })
        
    except Exception as e:
        return jsonify({
            "status": "error",
            "message": str(e)
        }), 500
    finally:
        close_db_connection(conn)


@app.route("/api/reset-table", methods=["POST"])
def reset_table():
    """Reset salary_data table"""
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("TRUNCATE TABLE salary_data")
        conn.commit()
        cursor.close()
        
        return jsonify({
            "status": "success",
            "message": "ลบข้อมูลและ reset ID เรียบร้อยแล้ว"
        })
        
    except Exception as e:
        return jsonify({
            "status": "error",
            "message": str(e)
        }), 500
    finally:
        close_db_connection(conn)


# ============================================================================
#                          FILE UPLOAD & PROCESSING
# ============================================================================

def save_to_mysql(df: pd.DataFrame) -> int:
    """Save DataFrame to MySQL"""
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        df.columns = [clean_column_name(c) for c in df.columns]
        
        # Create table if not exists
        cursor.execute("SHOW TABLES LIKE 'salary_data'")
        if not cursor.fetchone():
            col_defs = []
            for c in df.columns:
                clean_col = clean_column_name(c)
                
                if clean_col in ['name', 'employee', 'employee_type']:
                    col_defs.append(f"`{clean_col}` VARCHAR(255)")
                elif clean_col in ['cid', 'bank_account']:
                    col_defs.append(f"`{clean_col}` VARCHAR(20)")  # ✅ เปลี่ยนเป็น VARCHAR
                elif clean_col == 'year':
                    col_defs.append(f"`{clean_col}` INT")
                elif clean_col == 'month':
                    col_defs.append(f"`{clean_col}` TINYINT")
                elif clean_col in NUMERIC_COLUMNS:
                    col_defs.append(f"`{clean_col}` DECIMAL(15,2)")
                else:
                    col_defs.append(f"`{clean_col}` TEXT")
            
            cols_sql = ", ".join(col_defs)
            cursor.execute(f"""
                CREATE TABLE salary_data (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    {cols_sql},
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)
        
        saved = updated = failed = 0
        
        for idx, row in df.iterrows():
            try:
                processed_row = {}
                
                for col in df.columns:
                    clean_col = clean_column_name(col)
                    value = row[col]
                    
                    if clean_col == 'cid':
                        processed_row[clean_col] = clean_cid(value)
                    elif clean_col == 'year':
                        processed_row[clean_col] = clean_year(value)
                    elif clean_col == 'month':
                        processed_row[clean_col] = convert_month_to_number(value)
                    elif clean_col == 'bank_account':
                        processed_row[clean_col] = clean_bank_account(value)
                    elif clean_col in NUMERIC_COLUMNS:
                        processed_row[clean_col] = convert_to_decimal(value)
                    else:
                        v = clean_value(value)
                        processed_row[clean_col] = v if v != '' else None
                
                # เช็ค duplicate
                cid = processed_row.get('cid')
                month = processed_row.get('month')
                year = processed_row.get('year')
                
                exists = None
                
                if cid and month is not None and year is not None:
                    cursor.execute(
                        "SELECT id FROM salary_data WHERE cid = %s AND month = %s AND year = %s",
                        (cid, month, year)
                    )
                    exists = cursor.fetchone()
                
                columns = [clean_column_name(c) for c in df.columns]
                values = [processed_row.get(clean_column_name(c)) for c in df.columns]
                
                if exists:
                    set_clause = ", ".join([f"`{c}`=%s" for c in columns])
                    cursor.execute(
                        f"UPDATE salary_data SET {set_clause} WHERE id=%s",
                        values + [exists[0]]
                    )
                    updated += 1
                else:
                    cols_str = ", ".join([f"`{c}`" for c in columns])
                    placeholders = ", ".join(["%s"] * len(columns))
                    cursor.execute(
                        f"INSERT INTO salary_data ({cols_str}) VALUES ({placeholders})",
                        values
                    )
                    saved += 1
                    
            except Exception as e:
                print(f"Error processing row {idx}: {e}")
                import traceback
                traceback.print_exc()
                failed += 1
                continue

        conn.commit()
        cursor.close()

        print(f"✅ Saved: {saved}, Updated: {updated}, Failed: {failed}")
        return saved + updated
        
    finally:
        close_db_connection(conn)


@app.route("/upload", methods=["POST"])
def upload_api():
    """Upload and process Excel file"""
    if 'file' not in request.files:
        return jsonify({"error": "No file"}), 400
    
    if 'month' not in request.form or 'year' not in request.form:
        return jsonify({"error": "กรุณาเลือกเดือนและปี"}), 400
    
    file = request.files['file']
    selected_month = request.form['month']
    selected_year = request.form['year']
    
    selected_month_num = convert_month_to_number(selected_month)
    if selected_month_num is None:
        return jsonify({"error": "รูปแบบเดือนไม่ถูกต้อง"}), 400
    
    path = os.path.join(UPLOAD_FOLDER, file.filename)
    file.save(path)
    
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        dfs = []
        skipped = []
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            rows = list(sheet.iter_rows(values_only=True))
            
            if not rows:
                skipped.append(sheet_name)
                continue
            
            # Process header
            raw_header = []
            for h in rows[0]:
                if h is None or str(h).strip().lower() in ["", "nan", "none", "#ref!"]:
                    raw_header.append("col_empty")
                else:
                    raw_header.append(str(h).strip())
            
            data_rows = rows[1:]
            
            # Remove footer rows
            cut = None
            for i, r in enumerate(data_rows):
                if any(isinstance(c, str) and "รายละเอียดเพิ่มเติม" in c for c in r if c):
                    cut = i
                    break
            if cut is not None:
                data_rows = data_rows[:cut]
            
            # Filter empty rows
            filtered = []
            for r in data_rows:
                if r and not all(c in [None, "", " "] for c in r):
                    filtered.append(r)
            
            if not filtered:
                skipped.append(sheet_name)
                continue
            
            # Create DataFrame
            max_cols = max(len(raw_header), max(len(r) for r in filtered))
            header = raw_header + [f"col_empty_{i}" for i in range(len(raw_header), max_cols)]
            
            fixed_rows = [list(r) + ['']*(max_cols - len(r)) for r in filtered]
            df = pd.DataFrame(fixed_rows, columns=header)
            
            # Remove excluded columns
            df = df[[c for c in df.columns if c not in EXCLUDE_COLUMNS]]
            
            # ✅ Map column names และรวมคอลัมน์ที่ซ้ำกัน
            mapped_cols = []
            for c in df.columns:
                key = str(c).strip()
                if key.lower() in ["", "nan", "none"]:
                    key = "col_empty"
                mapped_cols.append(COLUMN_MAP.get(key, key))
            
            df.columns = mapped_cols
            
            # 🔧 รวมคอลัมน์ที่มีชื่อซ้ำกัน (แทนการแยกเป็น _1, _2)
            duplicate_cols = {}
            for col in df.columns:
                if col not in duplicate_cols:
                    duplicate_cols[col] = []
                duplicate_cols[col].append(col)
            
            # สร้าง DataFrame ใหม่โดยรวมค่าของคอลัมน์ซ้ำ
            merged_df = pd.DataFrame()
            processed_cols = set()
            
            for col in df.columns:
                if col in processed_cols:
                    continue
                    
                # หาคอลัมน์ทั้งหมดที่ชื่อเดียวกัน
                same_name_cols = [c for i, c in enumerate(df.columns) if df.columns[i] == col]
                
                if len(same_name_cols) > 1:
                    # ✅ รวมค่าจากหลายคอลัมน์ (เลือกค่าที่ไม่ใช่ null/empty)
                    combined_values = []
                    for idx in range(len(df)):
                        values = []
                        for sc in same_name_cols:
                            val = df.iloc[idx][sc]
                            if val not in [None, '', ' '] and not pd.isna(val):
                                # ถ้าเป็นตัวเลข ให้รวมกัน
                                if col in NUMERIC_COLUMNS:
                                    try:
                                        values.append(convert_to_decimal(val) or 0)
                                    except:
                                        pass
                                else:
                                    values.append(val)
                        
                        # สำหรับ numeric columns ให้รวมค่า
                        if col in NUMERIC_COLUMNS and values:
                            combined_values.append(sum(values))
                        # สำหรับ text columns ให้เอาค่าแรกที่ไม่ว่าง
                        elif values:
                            combined_values.append(values[0])
                        else:
                            combined_values.append(None)
                    
                    merged_df[col] = combined_values
                else:
                    merged_df[col] = df[col]
                
                processed_cols.add(col)
            
            df = merged_df
            
            # Clean column names
            df.columns = [clean_column_name(c) for c in df.columns]
            
            # Remove unwanted columns
            columns_to_drop = ['order_no', 'row_no']
            for col in columns_to_drop:
                if col in df.columns:
                    df = df.drop(columns=[col])
            
            # Add default CID if missing
            if 'cid' not in df.columns:
                df['cid'] = [f"{sheet_name}_{i}" for i in range(len(df))]
            
            # Add metadata
            df['employee'] = sheet_name
            df['month'] = selected_month_num
            df['year'] = selected_year
            
            dfs.append(df)
        
        if not dfs:
            return jsonify({
                "error": "No usable sheets",
                "skipped": skipped
            }), 400
        
        # Combine all sheets
        final_df = pd.concat(dfs, ignore_index=True)
        
        # Save to database
        saved = save_to_mysql(final_df)
        
        return jsonify({
            "status": "success",
            "processed_sheets": [d['employee'].iloc[0] for d in dfs],
            "skipped_sheets": skipped,
            "rows": len(final_df),
            "saved": saved,
            "selected_month": selected_month_num,
            "selected_year": selected_year
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 400
        
    finally:
        try:
            os.remove(path)
        except:
            pass

# ============================================================================
#                          MAIN
# ============================================================================

if __name__ == '__main__':
    app.run(debug=True, host="127.0.0.1", port=5000)